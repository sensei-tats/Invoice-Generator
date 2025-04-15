import tkinter as tk
from tkinter import messagebox
import pandas as pd
from fpdf import FPDF
import yagmail
import os
from dotenv import load_dotenv
from datetime import date, datetime
from openpyxl import load_workbook

# Load environment variables
load_dotenv()
EMAIL = os.getenv("SMTP_EMAIL")
PASSWORD = os.getenv("SMTP_PASSWORD")

# Paths
clients_file = "clients_projects.xlsx"
log_file = "invoice_log.xlsx"
invoice_dir = "invoices"
os.makedirs(invoice_dir, exist_ok=True)

# Read client data
df = pd.read_excel(clients_file)

# Load log
if os.path.exists(log_file):
    log_df = pd.read_excel(log_file)
else:
    log_df = pd.DataFrame(columns=["InvoiceID", "ClientName", "Email", "Amount", "DateSent", "DueDate", "Status"])

# Email setup
yag = yagmail.SMTP(EMAIL, PASSWORD)

# Generate and send invoices
for idx, row in df.iterrows():
    try:
        name = row['ClientName']
        email = row['Email']
        project = row['Project']
        hours = row['Hours']
        rate = row['RatePerHour']
        due = row['DueDate']
        total = hours * rate
        invoice_id = f"INV{1000 + idx}"
        date_sent = date.today().strftime("%Y-%m-%d")

        # Generate PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="INVOICE", ln=True, align='C')
        pdf.ln(10)
        pdf.cell(200, 10, txt=f"Invoice ID: {invoice_id}", ln=True)
        pdf.cell(200, 10, txt=f"Date: {date_sent}", ln=True)
        pdf.cell(200, 10, txt=f"Client: {name}", ln=True)
        pdf.cell(200, 10, txt=f"Project: {project}", ln=True)
        pdf.cell(200, 10, txt=f"Hours Worked: {hours}", ln=True)
        pdf.cell(200, 10, txt=f"Rate per Hour: ${rate}", ln=True)
        pdf.cell(200, 10, txt=f"Total: ${total}", ln=True)
        pdf.cell(200, 10, txt=f"Due Date: {due}", ln=True)

        pdf_path = os.path.join(invoice_dir, f"Invoice_{invoice_id}.pdf")
        pdf.output(pdf_path)

        # Send Email
        yag.send(
            to=email,
            subject=f"Invoice {invoice_id} for {project}",
            contents=f"Hi {name},\n\nPlease find attached the invoice for the project: {project}.\nTotal Amount: ${total}\nDue Date: {due}\n\nThank you!",
            attachments=pdf_path
        )

        # Log
        log_df.loc[len(log_df.index)] = [invoice_id, name, email, total, date_sent, due, "Sent"]

        print(f"Invoice sent to {name} at {email}")

    except Exception as e:
        print(f"Error processing {row['ClientName']}: {e}")

# Save updated log
log_df.to_excel(log_file, index=False)
